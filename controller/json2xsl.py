import json
from openpyxl import Workbook


def init_excel(tmpdict, ws,is_for_prove=True):
    ws.cell(1,1, "id")
    ws.cell(1,2, "obf_expr")
    ws.cell(1,3, "obf_leng")
    ws.cell(1,4, "obf_var")
    ws.cell(1,5, "result_expr")
    ws.cell(1,6, "result_leng")
    ws.cell(1,7, "result_var")
    ws.cell(1,8, "succ_time")
    # ws.cell(1,9, "fail_time")
    # ws.cell(1,10, "success")
    if is_for_prove:
        ws.cell(1, 11, "score1")
        ws.cell(1, 12, "score2")
        ws.cell(1, 13, "score3")
        ws.cell(1, 14, "time1")
        ws.cell(1, 15, "time2")
        ws.cell(1, 16, "time3")
        ws.cell(1, 17, "try_count")

def write_excel(tmpdict, id, ws,is_for_prove=True):
    ws.cell(id+2, 1, tmpdict["id"])
    for i,k in enumerate(tmpdict):
        if k == "id":
            continue
        ws.cell(id+2,i+2, tmpdict[k])
    ws.cell(id+2,1, tmpdict["id"])
    ws.cell(id+2,2, tmpdict["obf_expr"])
    ws.cell(id+2,3, tmpdict["obf_leng"])
    ws.cell(id+2,4, tmpdict["obf_var"])
    ws.cell(id+2,5, tmpdict["result_expr"])
    ws.cell(id+2,6, tmpdict["result_leng"])
    ws.cell(id+2,7, tmpdict["result_var"])
    ws.cell(id+2,8, tmpdict["time"])
    ws.cell(id+2,9, tmpdict["module"])
    # if tmpdict["module"] == "xyntia":
    #     ws.cell(id+2,10, tmpdict["success"])
    if "score_1" in tmpdict:
        ws.cell(id + 2, 11, tmpdict["score_1"])
        ws.cell(id + 2, 12, tmpdict["score_2"])
        ws.cell(id + 2, 13, tmpdict["score_3"])
        ws.cell(id + 2, 14, tmpdict["time_1"])
        ws.cell(id + 2, 15, tmpdict["time_2"])
        ws.cell(id + 2, 16, tmpdict["time_3"])
        ws.cell(id + 2, 17, tmpdict["try_count"])


# a = "/home/plas/work/PLASynth/controller/result/plasynth_timeout_10_qsynth_1028.json"
# a = "/home/plas/work/PLASynth/controller/mbablast_compare/plasynth_luby_qsynth_timeout_xyntia5_msynth1330(score20)_simp15_0313.json"
a = "/home/plas/work/PLASynth/controller/mbablast_compare/qsynth_old_plasynth.json"
write_wb = Workbook()
write_ws = write_wb.active
tmpdict = {}
succ_list = []
with open(a) as simp:
    simtime = 0.0
    syntime = 0.0
    time = 0.0
    time_fail = 0.0
    obf_leng = 0
    result_leng = 0
    obf_var = 0
    result_var = 0
    cnt = 0
    cnt_fail = 0
    count = 0
    msynth = json.load(simp)

    for i in msynth:
        try:
            tmp = i
            # if tmp["result_expr"] =="FAIL":
            #     time_fail += tmp["time"]
            #     tmp["fail_time"] = tmp["time"]
            #     cnt_fail += 1

            time += tmp["time"]
            obf_leng += tmp["obf_leng"]
            result_leng += tmp["result_leng"]
            obf_var += tmp["obf_var"]
            result_var += tmp["result_var"]
            cnt += 1
            succ_list.append(tmp["id"])

            id = int(tmp["id"])
            write_excel(tmp, id, write_ws)
            tmpdict = tmp
        except:
            print("error")

init_excel(tmpdict,write_ws)

outexcel = a.split("/")[-1][:-5]
write_wb.save(f"./{outexcel}.xlsx")