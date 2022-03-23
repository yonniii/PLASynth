from miasm.expression.expression import *
from openpyxl import Workbook
import re

def init_excel(tmpdict, ws, is_for_prove=True):
    ws.cell(1, 1, "id")
    ws.cell(1, 2, "obf_expr")
    ws.cell(1, 3, "obf_leng")
    ws.cell(1, 4, "obf_var")
    ws.cell(1, 5, "result_expr")
    ws.cell(1, 6, "result_leng")
    ws.cell(1, 7, "result_var")
    ws.cell(1, 8, "succ_time")
    # ws.cell(1,9, "fail_time")
    # ws.cell(1,10, "success")
    if is_for_prove:
        ws.cell(1, 11, "score1")
        ws.cell(1, 12, "score2")
        ws.cell(1, 13, "score3")
        ws.cell(1, 14, "time1")
        ws.cell(1, 15, "time2")
        ws.cell(1, 16, "time3")
        ws.cell(1, 17, "try_count")


def write_excel(obf, result, time, id, ws, is_for_prove=True):

    ws.cell(id + 2, 1, id-1)
    ws.cell(id + 2, 2, str(obf))
    ws.cell(id + 2, 3, obf.length())
    ws.cell(id + 2, 4, "")
    ws.cell(id + 2, 5, str(result))
    ws.cell(id + 2, 6, result.length())
    ws.cell(id + 2, 7, "")
    ws.cell(id + 2, 8, time[1:-1])
    ws.cell(id + 2, 9, "")


# a = "/home/plas/work/PLASynth/controller/result/plasynth_timeout_10_qsynth_1028.json"
def insertInt(n):
    m = n.group()
    return "ExprInt(%s, size)" % m


def string2ExprOp(strings, size=32):
    a = ExprId('v0', size)
    b = ExprId('v1', size)
    c = ExprId('v2', size)
    d = ExprId('v3', size)
    e = ExprId('v4', size)
    f = ExprId('v0', size)
    x = ExprId('v0', size)
    y = ExprId('v1', size)
    z = ExprId('v2', size)
    t = ExprId('v3', size)
    const_1 = ExprInt(1, size)
    const_2 = ExprInt(2, size)

    s = strings.replace("UL", "")
    s = re.sub(r'\b\d+\b', insertInt, s)
    s = re.sub(r'\b0x\w+\b', insertInt, s)
    return eval(s)


a = "/home/plas/work/msynth_test/synthesis_module/mba-blast/dataset/dataset2_32bit.txt.truthtable.search.simplify.txt"

write_wb = Workbook()
write_ws = write_wb.active
tmpdict = {}
succ_list = []
with open(a) as simp:
    simtime = 0.0
    syntime = 0.0
    time = 0.0
    time_fail = 0.0
    obf_leng = 0
    result_leng = 0
    obf_var = 0
    result_var = 0
    cnt = 0
    cnt_fail = 0
    count = 0
    lines = simp.readlines()

    for id, i in enumerate(lines):
        try:
            tmp = i
            if i.startswith("complex"):
                continue
            # if tmp["result_expr"] =="FAIL":
            #     time_fail += tmp["time"]
            #     tmp["fail_time"] = tmp["time"]
            #     cnt_fail += 1

            splited = i.split(",")
            origin = string2ExprOp(splited[0])
            simped = string2ExprOp(splited[2])
            isTrue = splited[3]
            time = splited[4]

            write_excel(origin, simped, time, id, write_ws)
            tmpdict = tmp
        except:
            print("error")

init_excel(tmpdict, write_ws)

outexcel = a.split("/")[-1][:-5]
write_wb.save(f"./{outexcel}.xlsx")

