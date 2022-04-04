import time
import json
from synthesis_module.module_handlers.xyntia import Xyntia
from synthesis_module.module_handlers.msynth_synthesis import Msynth_synth
from synthesis_module.module_handlers.msynth_simplifier import Msynth_simp
from get_sample import get_sample
from openpyxl import Workbook
from sample.handler.parser_others import string2ExprOp_list


# modules = [Msynth_simp()]
# modules = [Msynth_synth()]
# modules = [Xyntia()]
modules = [Msynth_synth(), Xyntia(),Msynth_simp()]
# modules = [ Xyntia(),Msynth_synth(),Msynth_simp()]

# modules = [Msynth_synth(), Xyntia(), Msynth_simp()]
# modules_re = [Xyntia(), Msynth_simp()]


def init_excel(ws):
    ws.cell(1,1, "id")
    ws.cell(1,2, "obf_expr")
    ws.cell(1,3, "obf_leng")
    ws.cell(1,4, "obf_var")
    ws.cell(1,5, "result_expr")
    ws.cell(1,6, "result_leng")
    ws.cell(1,7, "result_var")
    ws.cell(1,8, "time")
    ws.cell(1,9, "module")
    ws.cell(1,10, "success")




def write_excel(tmpdict, id, ws):
    ws.cell(id+2, 1, tmpdict["id"])
    for i,k in enumerate(tmpdict):
        if k == "id":
            continue
        ws.cell(id+2,i+2, tmpdict[k])
    ws.cell(id+2,1, tmpdict["id"])
    ws.cell(id+2,2, tmpdict["obf_expr"])
    ws.cell(id+2,3, tmpdict["obf_leng"])
    ws.cell(id+2,4, tmpdict["obf_var"])
    ws.cell(id+2,5, tmpdict["result_expr"])
    ws.cell(id+2,6, tmpdict["result_leng"])
    ws.cell(id+2,7, tmpdict["result_var"])
    ws.cell(id+2,8, tmpdict["time"])
    ws.cell(id+2,9, tmpdict["module"])
    if tmpdict["module"] == "xyntia":
        ws.cell(id+2,10, tmpdict["success"])


def run_PLASynth(samples, outfile,outfile_excel=""):
    file = open(outfile, "w")
    file.write("[\n")
    write_wb = Workbook()
    write_ws = write_wb.active
    # succ_count = 0
    # semantic_count = 0
    for i,expr in enumerate(samples):
        # if i != 447 and i != 448:
        #     continue
        if (i!=0):
            file.write(",")
        print(i)
        simplify_result = []
        obf_leng = expr.length()
        # obf_leng = 3
        start_time = time.time()

        # run each synthesis module=
        #######################################################
        for m_num, module in enumerate(modules):
            # if m_num == 0 and expr.length()<150:
            #     continue
            simplified, semantic_eq = module.simplify(expr)
            result_dict = module.getDict_fromExpr(expr, simplified,semantic_eq)
            if result_dict:
                simplify_result.append(result_dict)
            # if m_num == 0:
            #########
            # if m_num == 0 and simplified:
            #     break
            # if module.get_module_name()=="msynth-simp" and simplified and result_dict["result_leng"]>15:
            # # if module.get_module_name()=="msynth-simp" and simplified :
            #     simplified_tmp = simplified
            #     for m_num_2, module_2 in enumerate(modules):
            #         # print(simplified)
            #         # print(module_2)
            #         simplified, semantic_eq = module_2.simplify_try2(simplified_tmp,timeout=1)
            #         result_dict = module_2.getDict_fromExpr(expr, simplified, semantic_eq)
            #         if result_dict:
            #             simplify_result.append(result_dict)
            # #########
            # if simplified == expr:
            #     result_dict["is_origin"] = "true"
            # else:
            #     result_dict["is_origin"] = "false"
            if m_num == 0 and result_dict["result_expr"] != "FAIL":
            # if result_dict["result_expr"] != "FAIL":
                break
        #######################################################

        # module = modules[0]
        endtime = round(time.time() - start_time, 2)
        # select most effective result
        simplify_result.sort(key=lambda x:x["result_leng"])
        tmpdict_first = simplify_result[0]
        tmpmodule = tmpdict_first["module"]

        # # is_success = True if simplified.split("\n")[4].split(":")[1].strip() == "yes" else False
        # is_success = simplified   ## msynth-synth
        # tmpdict_first["is_success"] = True if is_success else False
        # tmpdict_first["semantic_eq"] = semantic_eq
        #
        # if is_success:
        #     succ_count += 1
        #     if semantic_eq:
        #         semantic_count += 1

        # if tmp:
        #     score_list = tmp[0]
        #     time_list = tmp[1]
        #     try_count = tmp[2]
        #     tmpdict_first['score_1'] = score_list[0]
        #     tmpdict_first['score_2'] = score_list[1]
        #     tmpdict_first['score_3'] = score_list[2]
        #     tmpdict_first['time_1'] = time_list[0]
        #     tmpdict_first['time_2'] = time_list[1]
        #     tmpdict_first['time_3'] = time_list[2]
        #     tmpdict_first['try_count'] = try_count

        # if tmpmodule != "xyntia":
        #
        #     for tmpexpr in string2ExprOp_list([tmpdict_first["result_expr"]]):
        #         expr = tmpexpr
        #     simplify_result = []
        #     for m_num, module in enumerate(modules_re):
        #         if m_num == 0 and expr.length()<150:
        #             continue
        #         simplified = module.simplify(expr)
        #         result_dict = module.getDict_fromExpr(expr, simplified)
        #         if result_dict:
        #             simplify_result.append(result_dict)
        #         if m_num == 0 and simplified:
        #             break
        #
        #     simplify_result.sort(key=lambda x:x["result_leng"])
        #     tmpdict = simplify_result[0]
        #     if tmpdict["result_leng"] < tmpdict_first["result_leng"]:
        #         tmpdict_first["result_expr"] = tmpdict["result_expr"]
        #         tmpdict_first["result_leng"] = tmpdict["result_leng"]
        #         tmpdict_first["module"] = "%s, %s" % (tmpmodule, tmpdict["module"])



        tmpdict_first['time'] = endtime
        tmpdict_first["id"] = i
        tmpdict_first["obf_leng"] = obf_leng

        file.write(json.dumps(tmpdict_first, indent=2))
        # write_excel(tmpdict_first,i,write_ws)
    init_excel(write_ws)
    file.write("\n]")
    file.close()
    # print(f"success : {succ_count}")
    # print(f"semantic : {semantic_count}")
    # print(f"percentage : {(semantic_count/succ_count)*100} %")
    # write_wb.save(outfile_excel)

def run(synthesis_module_type, sample_type):
    # filename = "./mbablast_compare/%s_brute.json" % (sample_type)
    # filename = "./mbablast_compare/plasynth_luby_%s_timeout_xyntia5_msynth1360(score10)_simp15_0313.json" % (sample_type)
    filename = "./mbablast_compare/%s_old_plasynth.json" % (sample_type)
    # filename = "./result_0124_finaltest/%s_%s_try2.json" % (synthesis_module_type, sample_type)
    run_PLASynth(samples=samples,outfile=filename)

if __name__ == "__main__":
    # pass
    synthesis_module_type = "plasynth" # don't modify
    sample_type = "qsynth" # select sample type {diff, qsynth, tigress, other}

    samples = get_sample(sample_type)
    run(synthesis_module_type, sample_type)

